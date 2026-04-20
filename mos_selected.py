from config import bot, DATABASE_URL, ADMIN_ID

import io
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import asyncpg
from aiogram.types import Message, InputFile
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, InputMediaDocument
from states import UploadHTML_MOS
from aiogram.types.input_file import FSInputFile
# Припускаємо, що bot, ADMIN_ID, DATABASE_URL визначені в іншому місці

async def start_upload_process_mos(message: Message, state: FSMContext):
    await message.answer("Скидуйте HTML файл.")
    await state.set_state(UploadHTML_MOS.html_required)

async def handle_html_document_mos(message: Message, state: FSMContext):
    await receive_html_file_mos(message, state)  # Викликаємо існуючу функцію обробки

async def invalid_input_in_state_mos(message: Message):
    await message.answer("Будь ласка, надішліть HTML файл. Якщо хочете скасувати, натисніть ⬅ Назад")

async def cancel_upload_mos(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Завантаження скасовано.")

# Модифікація receive_html_file для роботи зі станом:
# (Додати перевірки, і якщо не HTML, не очищати стан, а нагадувати)

# Оновлена версія receive_html_file (з попередньої)
async def receive_html_file_mos(message: Message, state: FSMContext):
    document = message.document

    # Ця перевірка вже зроблена фільтром, але на всяк випадок залишаємо
    # (якщо функцію викличуть без фільтрів)
    if not document or not document.file_name.lower().endswith('.html'):
        await message.answer("Будь ласка, надішліть файл у форматі <b>.html</b>", parse_mode="HTML")
        return  # залишаємося в стані

    status_msg = await message.answer(f"Обробляю файл: <b>{document.file_name}</b>...", parse_mode="HTML")

    try:
        # Завантаження та парсинг HTML
        file_io = io.BytesIO()
        await bot.download(document, file_io)
        file_io.seek(0)
        html_content = file_io.read().decode('utf-8', errors='ignore')

        soup = BeautifulSoup(html_content, "html.parser")
        qtext_elements = soup.find_all(class_="qtext")
        right_answer_elements = soup.find_all(class_="rightanswer")

        records = []
        for q_el, a_el in zip(qtext_elements, right_answer_elements):
            question = ' '.join(q_el.get_text().split())
            answer_raw = ' '.join(a_el.get_text().split())
            answer = answer_raw.split(':', 1)[-1].strip() if ':' in answer_raw else answer_raw.strip()

            if question and answer:
                records.append((question, answer))

        if not records:
            await status_msg.edit_text("⚠️ У файлі не знайдено валідних пар питання-відповідь.")
            await state.clear()
            return

        # Створення XLSX
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Питання та відповіді"

        # Заголовки зі стилем
        sheet["A1"] = "Питання"
        sheet["B1"] = "Відповідь"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4CAF50")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, (q, a) in enumerate(records, start=2):
            sheet[f"A{i}"] = q
            sheet[f"B{i}"] = a

        sheet.column_dimensions['A'].width = 60
        sheet.column_dimensions['B'].width = 60

        # Хто додав
        added_by = f"@{message.from_user.username}" if message.from_user.username else f"{message.from_user.first_name} ({message.from_user.id})"
        username_part = added_by.lstrip('@')  # видаляємо @ на початку, якщо є

        # Додатково: замінимо всі недопустимі для імені файлу символи на _
        safe_username = "".join(c if c.isalnum() or c in "_-" else "_" for c in username_part)
        # Збереження в пам'ять
        xlsx_io = io.BytesIO()
        workbook.save(xlsx_io)
        xlsx_io.seek(0)

        # 1. XLSX файл (згенерований)
        # XLSX файл
        xlsx_file = BufferedInputFile(
            xlsx_io.read(),
            filename=f"MOS_{safe_username}.xlsx"
        )

        # Оригінальний HTML файл (завантажуємо повторно)
        html_io = io.BytesIO()
        await bot.download(document, html_io)
        html_io.seek(0)

        html_file = BufferedInputFile(
            html_io.read(),
            filename=document.file_name
        )

        # Спільний caption (буде тільки у першому файлі)
        caption = (
            f"Новий тест додано користувачем: {added_by}\n"
            f"Оригінальний файл: <b>{document.file_name}</b>\n"
            f"Готовий XLSX: <b>MOS_{safe_username}.xlsx</b>\n"
            f"Питань: <b>{len(records)}</b>"
        )

        admin_ids = ADMIN_ID if isinstance(ADMIN_ID, list) else [ADMIN_ID]

        send_success = False
        for admin_id in admin_ids:
            try:
                media_group = [
                    InputMediaDocument(
                        media=html_file
                    ),
                    InputMediaDocument(
                        media=xlsx_file,
                        caption=caption,
                        parse_mode="HTML"
                        # caption тут ігнорується, тому не вказуємо
                    )
                ]

                await bot.send_media_group(
                    chat_id=admin_id,
                    media=media_group
                )
                send_success = True

            except Exception as e:
                print(f"Не вдалося надіслати альбом адміну {admin_id}: {e}")

        if not send_success:
            await status_msg.edit_text("Файл оброблено, але не вдалося надіслати жодному адміну.")

        # Збереження в БД
        records_with_added = [(q, a, added_by) for q, a in records]

        conn = await asyncpg.connect(DATABASE_URL)
        try:
            await conn.execute(f'''
                CREATE TABLE IF NOT EXISTS "MOS" (
                    id SERIAL PRIMARY KEY,
                    question TEXT NOT NULL,
                    answer TEXT NOT NULL,
                    added_by TEXT NOT NULL,
                    UNIQUE(question)
                );
            ''')

            await conn.executemany(f'''
                INSERT INTO "MOS" (question, answer, added_by)
                VALUES ($1, $2, $3)
                ON CONFLICT (question) DO UPDATE
                SET answer = EXCLUDED.answer,
                    added_by = EXCLUDED.added_by;
            ''', records_with_added)

            await status_msg.edit_text(
                f"✅ Успішно додано!\n\n"
                f"Файл: <b>{document.file_name}</b>\n"
                f"КІ-4 дякує вам!\n",
                parse_mode="HTML"
            )
        except Exception as db_e:
            await status_msg.edit_text(f"❌ Помилка збереження в базу даних:\n{db_e}")
        finally:
            await conn.close()

    except Exception as general_e:
        await status_msg.edit_text(f"❌ Критична помилка обробки файлу:\n{general_e}")
    finally:
        await state.clear()
